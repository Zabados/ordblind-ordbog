"""
audit_translations.py
----------------------
Produces a colour-coded Excel spreadsheet for reviewing DA→EN translation quality.

  Mode 1 — Heuristic only (default, no network):
      python scripts/audit_translations.py

  Mode 2 — dict.cc cross-check:
      python scripts/audit_translations.py --dictcc data/dictcc_da_en.txt

  How to get the dict.cc file (free, educational use allowed):
      1. https://www1.dict.cc/translation_file_request.php
      2. Select Danish ↔ English, download .txt → save as data/dictcc_da_en.txt
      3. Re-run with --dictcc flag

Output:
    translation_review.xlsx   — two-sheet Excel workbook:
        "Review"   — flagged entries (Danish text, wrong POS prefix, too long)
        "Missing"  — entries with no primary translation yet

Cell colour key:
    RED    — Danish characters or words found in English field (wrong)
    AMBER  — Structural issue: missing/spurious "to", overlong, same as headword
    GREY   — Cell is TODO (missing, not necessarily wrong)
    LAVENDER — dict.cc mismatch (none of our translations match dict.cc)
    No fill — looks clean
"""

import argparse
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ENTRIES_DIR = Path("entries/da-en")
OUT_XLSX    = Path("translation_review.xlsx")

# ── Colour constants ─────────────────────────────────────────────────────────
C_RED      = "FFCCCC"   # wrong text (Danish chars/words in English field)
C_AMBER    = "FFE0A0"   # structural issue
C_GREY     = "D9D9D9"   # TODO / missing
C_LAVENDER = "E8D5F5"   # dict.cc mismatch
C_HDR_BLUE = "2F5496"   # header row — Review sheet
C_HDR_GRN  = "375623"   # header row — Missing sheet
C_ROW_ALT  = "F2F2F2"   # alternating row tint
C_TAG_OK   = "D9EAD3"   # tags present — light green
C_TAG_MISS = "D9D9D9"   # tags missing — grey (same as TODO)

DANISH_STANDALONE = re.compile(
    r"\b(og|eller|ikke|som|der|til|fra|med|på|af|om|har|var|vil|kan|skal|"
    r"bliver|være|nogen|noget|dette|disse|hvad|hvem|hvor|når|mens|inden|"
    r"siden|efter|under|over|samt)\b",
    re.IGNORECASE,
)
DANISH_CHARS = re.compile(r"[æøåÆØÅ]")

LOANWORDS = frozenset(
    "januar februar marts april maj juni juli august september oktober "
    "november december virus internet email computer telefon hospital musik "
    "sport hotel menu bonus krise tema zone fase pilot test form standard "
    "normal basis status niveau system program projekt proces produkt "
    "service team club bus taxi film format signal data".split()
)

FREQ_ORDER = {"core": 0, "common": 1, "general": 2, "rare": 3, "TODO": 4}


# ── Helpers ──────────────────────────────────────────────────────────────────

def read_field(text: str, field: str) -> str:
    m = re.search(rf"^\s*{re.escape(field)}:\s*(.+)$", text, re.MULTILINE)
    return m.group(1).strip() if m else ""


def read_tags(text: str) -> list[str]:
    m = re.search(r'^tags:\n((?:[ \t]+-[^\n]+\n?)+)', text, re.MULTILINE)
    if not m:
        return []
    return [
        re.sub(r'^\s*-\s*', '', ln).strip()
        for ln in m.group(1).splitlines()
        if ln.strip().startswith('-') and ln.strip().lstrip('- ') not in {'TODO', 'SKIP', ''}
    ]


def read_secondaries(text: str) -> list[str]:
    m = re.search(r"secondary_translations:(.*?)```", text, re.DOTALL)
    if not m:
        return []
    lines = []
    for raw in m.group(1).split("\n"):
        s = raw.strip()
        if s.startswith("- ") and s not in ("- SKIP", "- TODO"):
            lines.append(s[2:].strip())
    return lines


def normalise(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"^(a |an |the )", "", s)
    s = re.sub(r"[^\w\s]", "", s)
    return s.strip()


def cell_color(value: str, pos: str, headword: str) -> str:
    """Return hex fill colour for one translation cell, or '' for clean."""
    if not value or value == "SKIP":
        return ""
    if value == "TODO":
        return C_GREY
    if DANISH_CHARS.search(value):
        return C_RED
    if DANISH_STANDALONE.search(value):
        return C_RED
    if len(value) > 70:
        return C_AMBER
    if pos == "verb" and not value.lower().startswith("to "):
        return C_AMBER
    if pos != "verb" and value.lower().startswith("to "):
        return C_AMBER
    norm, hw = normalise(value), normalise(headword)
    if norm and hw and norm == hw and hw not in LOANWORDS:
        return C_AMBER
    return ""


# ── dict.cc loader ───────────────────────────────────────────────────────────

def load_dictcc(path: Path) -> dict[str, list[str]]:
    lookup: dict[str, list[str]] = {}
    if not path.exists():
        raise FileNotFoundError(f"dict.cc file not found: {path}")
    with open(path, encoding="utf-8", errors="replace") as fh:
        for line in fh:
            if line.startswith("#") or not line.strip():
                continue
            parts = line.rstrip("\n").split("\t")
            if len(parts) < 2:
                continue
            col0, col1 = parts[0].strip(), parts[1].strip()
            if DANISH_CHARS.search(col1) or _looks_danish(col1):
                da_term, en_term = col1, col0
            else:
                da_term, en_term = col0, col1
            da_clean = re.sub(r"[\[{(][^\]})]*[\]})][^\s]*", "", da_term).strip().lower()
            en_clean = re.sub(r"[\[{(][^\]})]*[\]})][^\s]*", "", en_term).strip()
            if da_clean and en_clean:
                lookup.setdefault(da_clean, []).append(en_clean)
    return lookup


def _looks_danish(s: str) -> bool:
    markers = {"og", "eller", "ikke", "som", "der", "til", "fra", "med", "på", "af"}
    return sum(1 for w in s.lower().split() if w in markers) >= 2


# ── Per-entry analysis ───────────────────────────────────────────────────────

def analyse_entry(path: Path, dictcc: dict[str, list[str]] | None) -> dict | None:
    text = path.read_text(encoding="utf-8")
    headword       = read_field(text, "headword")
    pos            = read_field(text, "pos")
    frequency_tier = read_field(text, "frequency_tier")
    primary        = read_field(text, "primary_translation")
    secondaries    = read_secondaries(text)
    tags           = read_tags(text)

    is_todo = (primary == "TODO")

    prim_color = cell_color(primary, pos, headword)
    sec_colors = [cell_color(s, pos, headword) for s in secondaries]

    has_issue = bool(prim_color) or any(sec_colors)

    # dict.cc cross-check
    dictcc_note  = ""
    dictcc_color = ""
    if dictcc is not None and not is_todo:
        dc_vals = dictcc.get(headword.lower(), [])
        if not dc_vals:
            dictcc_note  = "not in dict.cc"
            dictcc_color = C_LAVENDER
        else:
            all_vals  = ([primary] if primary not in ("TODO", "SKIP") else []) + secondaries
            our_norms = {normalise(v) for v in all_vals}
            dc_norms  = {normalise(v) for v in dc_vals}
            if not (our_norms & dc_norms):
                dictcc_note  = " / ".join(dc_vals[:5])
                dictcc_color = C_LAVENDER
                has_issue = True

    if not has_issue and not is_todo:
        return None  # clean entry

    # Build human-readable notes
    notes_parts: list[str] = []
    if prim_color == C_RED:
        notes_parts.append("DA text in primary")
    elif prim_color == C_AMBER:
        if pos == "verb" and primary and not primary.lower().startswith("to "):
            notes_parts.append("verb: add 'to'")
        elif pos != "verb" and primary and primary.lower().startswith("to "):
            notes_parts.append("non-verb: remove 'to'")
        elif primary and len(primary) > 70:
            notes_parts.append("primary > 70 chars")
        else:
            notes_parts.append("check primary")
    for i, (s, c) in enumerate(zip(secondaries, sec_colors), 1):
        if c == C_RED:
            notes_parts.append(f"DA text in sec.{i}")
        elif c == C_AMBER:
            if len(s) > 70:
                notes_parts.append(f"sec.{i} > 70 chars")
            elif pos == "verb" and not s.lower().startswith("to "):
                notes_parts.append(f"sec.{i}: add 'to'")
            elif pos != "verb" and s.lower().startswith("to "):
                notes_parts.append(f"sec.{i}: remove 'to'")
            else:
                notes_parts.append(f"check sec.{i}")
    if dictcc_note:
        label = "not in dict.cc" if dictcc_note == "not in dict.cc" else f"dict.cc: {dictcc_note}"
        notes_parts.append(label)

    return {
        "headword":       headword,
        "pos":            pos,
        "frequency_tier": frequency_tier,
        "primary":        primary,
        "prim_color":     prim_color,
        "secondaries":    secondaries,
        "sec_colors":     sec_colors,
        "dictcc_note":    dictcc_note,
        "dictcc_color":   dictcc_color,
        "notes":          "; ".join(notes_parts),
        "is_todo":        is_todo,
        "tags":           tags,
        "file":           path.name,
    }


# ── Excel helpers ────────────────────────────────────────────────────────────

def make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def set_col_width(ws, col: int, width: float):
    ws.column_dimensions[get_column_letter(col)].width = width


def write_header(ws, headers: list[str], hdr_color: str):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = Font(bold=True, color="FFFFFF", name="Verdana", size=10)
        c.fill      = make_fill(hdr_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28


def write_cell(ws, row: int, col: int, value: str, fill_hex: str = "",
               bold: bool = False, wrap: bool = True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Verdana", size=10, bold=bold)
    c.alignment = Alignment(vertical="top", wrap_text=wrap)
    if fill_hex:
        c.fill = make_fill(fill_hex)
    return c


# ── Build workbook ───────────────────────────────────────────────────────────

def build_xlsx(review: list[dict], missing: list[dict], dictcc_mode: bool):
    wb = Workbook()

    # ── Sheet 1: Review ───────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Review"

    MAX_SEC = max((len(r["secondaries"]) for r in review), default=0)
    MAX_SEC = max(MAX_SEC, 3)

    sec_headers = [f"Secondary {i}" for i in range(1, MAX_SEC + 1)]
    headers1 = ["DA word", "POS", "Freq", "EN primary"] + sec_headers + ["Tags", "Notes", "File"]
    write_header(ws1, headers1, C_HDR_BLUE)
    ws1.freeze_panes = "A2"

    set_col_width(ws1, 1, 18)
    set_col_width(ws1, 2, 11)
    set_col_width(ws1, 3, 9)
    set_col_width(ws1, 4, 36)
    for i in range(MAX_SEC):
        set_col_width(ws1, 5 + i, 36)
    tags_col  = 5 + MAX_SEC
    notes_col = tags_col + 1
    file_col  = notes_col + 1
    set_col_width(ws1, tags_col,  28)
    set_col_width(ws1, notes_col, 40)
    set_col_width(ws1, file_col,  26)

    for row_idx, r in enumerate(review, 2):
        alt = C_ROW_ALT if row_idx % 2 == 0 else ""

        write_cell(ws1, row_idx, 1, r["headword"],        alt,                 bold=True, wrap=False)
        write_cell(ws1, row_idx, 2, r["pos"],             alt,                 wrap=False)
        write_cell(ws1, row_idx, 3, r["frequency_tier"],  alt,                 wrap=False)
        write_cell(ws1, row_idx, 4, r["primary"],         r["prim_color"] or alt)

        for i in range(MAX_SEC):
            val   = r["secondaries"][i] if i < len(r["secondaries"]) else ""
            color = r["sec_colors"][i]  if i < len(r["sec_colors"])  else ""
            write_cell(ws1, row_idx, 5 + i, val, color or alt)

        tag_str   = ", ".join(r["tags"]) if r.get("tags") else ""
        tag_color = C_TAG_OK if tag_str else C_TAG_MISS
        write_cell(ws1, row_idx, tags_col,  tag_str,       tag_color,           wrap=False)
        write_cell(ws1, row_idx, notes_col, r["notes"],   r["dictcc_color"] or alt, wrap=True)
        write_cell(ws1, row_idx, file_col,  r["file"],    alt,                 wrap=False)

    # ── Sheet 2: Missing ─────────────────────────────────────────────────
    ws2 = wb.create_sheet(f"Missing ({len(missing)})")

    MAX_SEC2 = max((len(r["secondaries"]) for r in missing), default=0)
    MAX_SEC2 = max(MAX_SEC2, 2)
    sec_headers2 = [f"Secondary {i}" for i in range(1, MAX_SEC2 + 1)]
    headers2 = ["DA word", "POS", "Freq", "EN primary (TODO)"] + sec_headers2 + ["Tags", "File"]
    write_header(ws2, headers2, C_HDR_GRN)
    ws2.freeze_panes = "A2"

    set_col_width(ws2, 1, 18)
    set_col_width(ws2, 2, 11)
    set_col_width(ws2, 3, 9)
    set_col_width(ws2, 4, 36)
    for i in range(MAX_SEC2):
        set_col_width(ws2, 5 + i, 36)
    tags_col2 = 5 + MAX_SEC2
    file_col2 = tags_col2 + 1
    set_col_width(ws2, tags_col2, 28)
    set_col_width(ws2, file_col2, 26)

    missing_sorted = sorted(
        missing,
        key=lambda r: (FREQ_ORDER.get(r["frequency_tier"], 4), r["headword"])
    )

    for row_idx, r in enumerate(missing_sorted, 2):
        alt = C_ROW_ALT if row_idx % 2 == 0 else ""
        write_cell(ws2, row_idx, 1, r["headword"],        alt, bold=True, wrap=False)
        write_cell(ws2, row_idx, 2, r["pos"],             alt, wrap=False)
        write_cell(ws2, row_idx, 3, r["frequency_tier"],  alt, wrap=False)
        write_cell(ws2, row_idx, 4, "TODO",               C_GREY)
        for i in range(MAX_SEC2):
            val = r["secondaries"][i] if i < len(r["secondaries"]) else ""
            write_cell(ws2, row_idx, 5 + i, val, alt)
        tag_str   = ", ".join(r["tags"]) if r.get("tags") else ""
        tag_color = C_TAG_OK if tag_str else C_TAG_MISS
        write_cell(ws2, row_idx, tags_col2, tag_str, tag_color, wrap=False)
        write_cell(ws2, row_idx, file_col2, r["file"], alt, wrap=False)

    # ── Sheet 3: Legend ───────────────────────────────────────────────────
    ws3 = wb.create_sheet("Legend")
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 60
    legend_rows = [
        ("Colour",    "Meaning"),
        ("RED",       "Danish characters (æ/ø/å) or Danish words in English field — almost certainly wrong"),
        ("AMBER",     "Structural issue: 'to' prefix wrong for POS, translation > 70 chars, or same as DA word"),
        ("GREY",      "Cell is TODO — translation or tags missing (not necessarily wrong)"),
        ("GREEN",     "Tags column: at least one thematic tag is present"),
        ("LAVENDER",  "dict.cc cross-check: none of our translations matched dict.cc results"),
        ("No fill",   "Looks clean by heuristics"),
    ]
    fill_map = {"RED": C_RED, "AMBER": C_AMBER, "GREY": C_GREY, "GREEN": C_TAG_OK, "LAVENDER": C_LAVENDER}
    for ri, (label, desc) in enumerate(legend_rows, 1):
        c1 = ws3.cell(row=ri, column=1, value=label)
        c2 = ws3.cell(row=ri, column=2, value=desc)
        if label in fill_map:
            c1.fill = make_fill(fill_map[label])
        c1.font = Font(name="Verdana", size=10, bold=(ri == 1))
        c2.font = Font(name="Verdana", size=10, bold=(ri == 1))
        c2.alignment = Alignment(wrap_text=True)
        ws3.row_dimensions[ri].height = 22

    wb.save(OUT_XLSX)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Produce colour-coded translation review spreadsheet."
    )
    parser.add_argument("--dictcc", metavar="FILE",
                        help="Path to dict.cc DA-EN vocabulary file (optional)")
    args = parser.parse_args()

    dictcc: dict[str, list[str]] | None = None
    if args.dictcc:
        print(f"Loading dict.cc file: {args.dictcc} …")
        dictcc = load_dictcc(Path(args.dictcc))
        print(f"  Loaded {len(dictcc):,} Danish headwords")

    entries = sorted(ENTRIES_DIR.glob("*.md"))
    print(f"Analysing {len(entries)} entries …")

    review:  list[dict] = []
    missing: list[dict] = []

    for path in entries:
        result = analyse_entry(path, dictcc)
        if result is None:
            continue
        if result["is_todo"] and not any(result["sec_colors"]) and not result["dictcc_note"]:
            missing.append(result)
        else:
            review.append(result)

    def severity(r: dict) -> int:
        colors = [r["prim_color"]] + r["sec_colors"] + [r["dictcc_color"]]
        if C_RED      in colors: return 0
        if C_AMBER    in colors: return 1
        if C_LAVENDER in colors: return 2
        return 3

    review.sort(key=lambda r: (severity(r), FREQ_ORDER.get(r["frequency_tier"], 4), r["headword"]))

    build_xlsx(review, missing, dictcc is not None)

    print(f"\nReview sheet  : {len(review)} entries flagged")
    print(f"Missing sheet : {len(missing)} entries need primary translation")
    print(f"\nWritten: {OUT_XLSX}")
    print("Open in Excel or LibreOffice Calc.")
    if not args.dictcc:
        print("\nFor dict.cc cross-check:")
        print("  1. https://www1.dict.cc/translation_file_request.php")
        print("  2. Save DA-EN file as data/dictcc_da_en.txt")
        print("  3. python scripts/audit_translations.py --dictcc data/dictcc_da_en.txt")


if __name__ == "__main__":
    main()
