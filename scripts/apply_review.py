"""
apply_review.py
---------------
Reads translation_review.xlsx (after you've edited it) and writes the
changes back to the corresponding entry .md files.

Usage:
    python scripts/apply_review.py [--dry-run]

What it does:
    - Reads both the "Review" and "Missing" sheets
    - For each row: compares EN primary + secondaries against the .md file
    - If anything changed, rewrites only those two fields in the .md file
    - Prints a summary of every change made

Safe to re-run: only writes files where something actually changed.
"""

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook

ENTRIES_DIR  = Path("entries/da-en")
REVIEW_XLSX  = Path("translation_review.xlsx")


# ── Helpers ──────────────────────────────────────────────────────────────────

def read_field(text: str, field: str) -> str:
    m = re.search(rf"^\s*{re.escape(field)}:\s*(.+)$", text, re.MULTILINE)
    return m.group(1).strip() if m else ""


def read_secondaries(text: str) -> list[str]:
    m = re.search(r"secondary_translations:(.*?)```", text, re.DOTALL)
    if not m:
        return []
    lines = []
    for raw in m.group(1).split("\n"):
        s = raw.strip()
        if s.startswith("- "):
            lines.append(s[2:].strip())
    return lines


def set_primary(text: str, new_value: str) -> str:
    return re.sub(
        r"(^\s*primary_translation:\s*)(.+)$",
        lambda m: m.group(1) + new_value,
        text,
        flags=re.MULTILINE,
        count=1,
    )


def set_secondaries(text: str, values: list[str]) -> str:
    if not values:
        new_block = "secondary_translations:\n  - SKIP\n"
    else:
        lines = "\n".join(f"  - {v}" for v in values)
        new_block = f"secondary_translations:\n{lines}\n"

    return re.sub(
        r"secondary_translations:.*?(?=\n```)",
        new_block.rstrip("\n"),
        text,
        flags=re.DOTALL,
        count=1,
    )


def cell_str(cell) -> str:
    """Return cell value as a clean string, or '' for None/whitespace."""
    v = cell.value
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s in ("TODO", "SKIP") else s


# ── Sheet reader ─────────────────────────────────────────────────────────────

def read_sheet(ws) -> list[dict]:
    """
    Read a Review or Missing sheet into a list of row dicts.
    Returns: [{file, primary, secondaries: [str, ...]}, ...]
    """
    headers = [c.value for c in ws[1]]

    # Locate key columns by header name
    def col_idx(name: str) -> int | None:
        for i, h in enumerate(headers):
            if h and name.lower() in str(h).lower():
                return i
        return None

    file_col    = col_idx("file")
    primary_col = col_idx("en primary")
    sec_cols    = sorted(
        [i for i, h in enumerate(headers) if h and str(h).lower().startswith("secondary")],
        key=lambda i: headers[i],
    )

    if file_col is None or primary_col is None:
        return []

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        file_val = cell_str(row[file_col])
        if not file_val:
            continue
        primary_val = cell_str(row[primary_col])
        sec_vals = [cell_str(row[i]) for i in sec_cols if cell_str(row[i])]

        rows.append({
            "file":        file_val,
            "primary":     primary_val,
            "secondaries": sec_vals,
        })
    return rows


# ── Apply changes ─────────────────────────────────────────────────────────────

def apply_row(row: dict, dry_run: bool) -> str | None:
    """
    Compare xlsx row against the .md file. If changed, write the file.
    Returns a description of the change, or None if nothing changed.
    """
    path = ENTRIES_DIR / row["file"]
    if not path.exists():
        return f"  SKIP (file not found): {row['file']}"

    text     = path.read_text(encoding="utf-8")
    old_prim = read_field(text, "primary_translation")
    old_secs = read_secondaries(text)

    new_prim = row["primary"] or old_prim   # blank cell = no change
    new_secs = row["secondaries"]

    # Normalise old secondaries for comparison (strip SKIP/TODO sentinels)
    old_secs_clean = [s for s in old_secs if s not in ("SKIP", "TODO")]

    changed_primary    = new_prim != old_prim and bool(row["primary"])
    changed_secondaries = new_secs != old_secs_clean and bool(row["secondaries"]) or (
        # User cleared all secondaries: old had real values, new is empty
        not new_secs and old_secs_clean
    )

    if not changed_primary and not changed_secondaries:
        return None

    # Build change description
    parts = []
    if changed_primary:
        parts.append(f"  primary:  '{old_prim}'  →  '{new_prim}'")
    if changed_secondaries:
        old_str = " | ".join(old_secs_clean) or "(none)"
        new_str = " | ".join(new_secs) or "(none)"
        parts.append(f"  secondaries: {old_str}  →  {new_str}")

    if not dry_run:
        updated = text
        if changed_primary:
            updated = set_primary(updated, new_prim)
        if changed_secondaries:
            updated = set_secondaries(updated, new_secs)
        path.write_text(updated, encoding="utf-8")

    return "\n".join(parts)


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Apply Excel review edits back to .md files.")
    parser.add_argument("--dry-run", action="store_true",
                        help="Show what would change without writing files")
    args = parser.parse_args()

    if not REVIEW_XLSX.exists():
        print(f"ERROR: {REVIEW_XLSX} not found. Run audit_translations.py first.")
        return

    print(f"Reading {REVIEW_XLSX} …")
    wb = load_workbook(REVIEW_XLSX, data_only=True)

    all_rows: list[dict] = []
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "legend":
            continue
        ws = wb[sheet_name]
        rows = read_sheet(ws)
        all_rows.extend(rows)
        print(f"  Sheet '{sheet_name}': {len(rows)} data rows")

    if args.dry_run:
        print("\n[DRY RUN — no files will be written]\n")

    changed = 0
    skipped = 0
    for row in all_rows:
        result = apply_row(row, args.dry_run)
        if result:
            if result.startswith("  SKIP"):
                print(result)
            else:
                changed += 1
                label = f"{row['file']:35s}"
                print(f"{label}")
                print(result)
        else:
            skipped += 1

    print(f"\n{'[DRY RUN] ' if args.dry_run else ''}Files {'would be ' if args.dry_run else ''}changed: {changed}")
    print(f"Unchanged: {skipped}")

    if changed and not args.dry_run:
        print("\nRe-render when ready:")
        print("  python scripts/render_sample.py --full --open")


if __name__ == "__main__":
    main()
